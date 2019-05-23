# -*- coding: utf-8 -*-
# COPYRIGHT (C) 2018-2019 GIG TECHNOLOGY NV
# ALL RIGHTS RESERVED.
#
# ALTHOUGH YOU MAY BE ABLE TO READ THE CONTENT OF THIS FILE, THIS FILE
# CONTAINS CONFIDENTIAL INFORMATION OF GIG TECHNOLOGY NV. YOU ARE NOT ALLOWED
# TO MODIFY, REPRODUCE, DISCLOSE, PUBLISH OR DISTRIBUTE ITS CONTENT,
# EMBED IT IN OTHER SOFTWARE, OR CREATE DERIVATIVE WORKS, UNLESS PRIOR
# WRITTEN PERMISSION IS OBTAINED FROM GIG TECHNOLOGY NV.
#
# THE COPYRIGHT NOTICE ABOVE DOES NOT EVIDENCE ANY ACTUAL OR INTENDED
# PUBLICATION OF SUCH SOURCE CODE.
#
# @@license_version:1.8@@

# ---------------------------
# resourse consumtion calculation
# (C) 2018 by Herby
#
# you only need to set this variables from bash
#
#--> IYO
#export CLIENT_ID="...."
#export CLIENT_SECRET="....."
#
# --> date for the report period
#export START=$(date -d '10/01/2018 00:00:00' +"%s")
#export END=$(date -d '10/30/2018 23:59:59' +"%s")
#
#--> excel file names
#export XLS_FILE="consumption_2018_10.xlsx"
#export XLS_FILE_TPL="consumption_tpl.xlsx"
#
# then run the script
# 1. to get your Accounts
#python3 export_all_in_zip.py list
#
# 2. to generate the XLSX
#python3 export_all_in_zip.py <account_id>
#
# and set the URL

URL='https://ch-lug-dc01-001.gig.tech/restmachine/cloudapi/'
URL='https://at-vie-dc01-001.gig.tech/restmachine/cloudapi/'


import os
from pathlib import Path
import capnp
import datetime
from os import environ
import sys
import shutil
from shutil import copyfile
import requests
import openpyxl
import zipfile

temp_zipfile = 'constemp.zip'

def getKey(item):
    return item[0]

def getJWT():
    iyourl = 'https://itsyou.online/v1/oauth/access_token'

    params = {'grant_type': 'client_credentials',
    	      'client_id': os.environ["CLIENT_ID"],
    		  'response_type': 'id_token',
    		  'client_secret': os.environ["CLIENT_SECRET"]}
    resp = requests.post(iyourl, params=params)

    if resp.status_code != 200:
    	print ('error in JWT')
    	sys.exit(2)

    return resp.text

def getZPFile():
    #construct get for consumtion ZipFile
    ZURL= URL+'accounts/getConsumption?accountId='+account+'&start='+start_date+'&end='+end_date
    r = requests.get(ZURL, headers=headers, stream=True)
    with open(temp_zipfile, 'wb') as f:
        shutil.copyfileobj(r.raw, f)
    return

def getCSInfo(actcs):
    params = {'cloudspaceId': actcs}
    resp = requests.post(URL+"cloudspaces/get", headers=headers,params=params)
    css=resp.json()
    return (css['status'],css['name'],css['description'],css['externalnetworkip'])

def getAccountInfo(accid):
    params = {'accountId': accid}
    resp = requests.post(URL+"accounts/get", headers=headers,params=params)
    ac=resp.json()
    return (ac['status'],ac['name'],ac['resourceLimits'])

def getAccounts():
    resp = requests.post(URL+"accounts/list", headers=headers)
    ac=resp.json()
    print ('"Accessable Accounts for given rights')
    print ('On G8: ',URL)
    print ('----------------------------------------')
    for a in ac:
        print ("ID:",a['id']," Name:",a['name'])
    return


# get environ
xlsfile = environ["XLS_FILE"]
xlstpl = environ["XLS_FILE_TPL"]
start_date = environ["START"]
end_date = environ["END"]

# JWT
JWT = getJWT()
if JWT == None:
	print ('No JWT')
	sys.exit(1)

AUTH = "Bearer " + JWT
headers = {'Authorization': AUTH}

if len(sys.argv) == 1:
        print (' no command ')
        print ('----------------------------------------')
        print (' usage .....')
        print (sys.argv[0]," list")
        print (sys.argv[0],"<AccountID>")
        print ('----------------------------------------')
        sys.exit(1)

if sys.argv[1] == "list":
    getAccounts()
    sys.exit(0)

# check arguments!!! #######################
#if sys.argv[1] :
#    sys.exit(1)

account = sys.argv[1]

#init excel file from template
copyfile(xlstpl, xlsfile)

# init capnp, and fetch if not here
capnp.remove_import_hook()
resourcemonitoring = "resourcemonitoring.capnp"
rc_file = Path(resourcemonitoring)
if not rc_file.exists():
    r = requests.get("https://git.gig.tech/openvcloud/openvcloud/raw/master/libs/CloudscalerLibcloud/CloudscalerLibcloud/schemas/resourcemonitoring.capnp?private_token=<access token>", stream=True)
    f = open(resourcemonitoring, 'w')
    f.write (r.text)
    f.close()
resources_capnp = capnp.load("resourcemonitoring.capnp")

#get the resourse consumption zip file
getZPFile()

wb = openpyxl.load_workbook(xlsfile)
ws = wb['a']

idx=8
ws['A'+str(idx)]= 'Account'
ws['B'+str(idx)]= 'date'
ws['C'+str(idx)]=  'Year'
ws['D'+str(idx)]= 'Month'
ws['E'+str(idx)]= 'Day'
ws['F'+str(idx)]= 'Hour'
ws['G'+str(idx)]= 'Cloud Space ID'
ws['H'+str(idx)]= 'Machine Count'
ws['I'+str(idx)]= 'Total Memory'
ws['J'+str(idx)]= 'Total VCPUs'
ws['K'+str(idx)]= 'Disk Size'
ws['L'+str(idx)]= 'Disk IOPS Read'
ws['M'+str(idx)]= 'Disk IOPS Write'
ws['N'+str(idx)]= 'NICs TX'
ws['O'+str(idx)]= 'NICs RX'
ws['P'+str(idx)]= 'Win'
ws['Q'+str(idx)]= 'Redhat'
ws['R'+str(idx)]= '----'

mi = []
ci = []

with zipfile.ZipFile(temp_zipfile) as zfile:
    for file in zfile.namelist():
        path_list = file.split(os.sep)
        account = path_list[0]
        year = path_list[1]
        month = path_list[2]
        day = path_list[3]
        hour = path_list[4]
        date = datetime.datetime(int(year),int(month),int(day),int(hour)).strftime('%s')
        account_obj = resources_capnp.Account.from_bytes(zfile.read(file))
        for i, cs in enumerate(account_obj.cloudspaces):
            cs_id = cs.cloudSpaceId
            machines = len(cs.machines)
            vcpus = 0
            mem = 0
            disksize = 0
            disk_iops_read = 0
            disk_iops_write = 0
            nics_tx = 0
            nics_rx = 0
            win = 0
            redhat = 0
            for machine in cs.machines:
                #if machine.status = "RUNNING"
                m_su = 0
                m_tu = 0
                m_nu = 0
                m_win = 0
                m_rh = 0
                if machine.id == 0:
                    continue
                vcpus += machine.vcpus
                mem += machine.mem
                #print ("cpuMinutes:",machine.cpuMinutes," status:",machine.status)
                #print (cs.cloudSpaceId,"  ",machine.id,"  ",machine.imageName)
                if machine.imageName.find('Windows') >=0:
                    win += 1
                    m_win = 1
                if machine.imageName.find('Redhat') >=0:
                    redhat += 1
                    m_rh = 1

                for disk in machine.disks:
                    disk_iops_read += disk.iopsRead
                    disk_iops_write += disk.iopsWrite
                    disksize += disk.size
                    m_su += disk.size
                    m_tu += (disk.iopsRead+disk.iopsWrite)
                for nic in machine.networks:
                    nics_tx += nic.tx
                    nics_rx += nic.rx
                    m_nu += (nic.tx+nic.rx)
                # list all used cloudspaces
                cc = cs.cloudSpaceId
                if not (cc in ci):
                    ci.append(cc)

                # list all machines
                mmm = (machine.id)
                m = [row[1] for row in mi]
                try:
                    midx = m.index(mmm)
                except Exception as e:
                    midx = -1
                if midx >= 0:
                    mi[midx][3] += machine.vcpus
                    mi[midx][4] += machine.mem
                    mi[midx][5] += m_su
                    mi[midx][6] += m_tu
                    mi[midx][7] += m_nu
                    mi[midx][8] += m_win
                    mi[midx][9] += m_rh
                else:
                    #print ("neu:",)
                    mmm = [cs.cloudSpaceId,machine.id,machine.imageName,machine.vcpus,machine.mem,m_su,m_tu,m_nu,m_win,m_rh]
                    mi.append(mmm)

            idx +=1

            ws['A'+str(idx)]=account
            ws['B'+str(idx)]=int(date)
            ws['C'+str(idx)]=year
            ws['D'+str(idx)]=month
            ws['E'+str(idx)]=day
            ws['F'+str(idx)]=hour
            ws['G'+str(idx)]=cs_id
            ws['H'+str(idx)]=machines
            ws['I'+str(idx)]=mem
            ws['J'+str(idx)]=vcpus
            ws['K'+str(idx)]= disksize
            ws['L'+str(idx)]= disk_iops_read
            ws['M'+str(idx)]= disk_iops_write
            ws['N'+str(idx)]= nics_tx
            ws['O'+str(idx)]= nics_rx
            ws['P'+str(idx)]= win
            ws['Q'+str(idx)]= 1
            ws['R'+str(idx)]= 1
            ws['S'+str(idx)]="=param!C3*I"+str(idx)
            ws['T'+str(idx)]="=param!C2*J"+str(idx)
            ws['U'+str(idx)]="=param!C4*K"+str(idx)
            ws['V'+str(idx)]="=param!C5*(L"+str(idx)+"+M"+str(idx)+")"
            ws['W'+str(idx)]="=param!C6*(N"+str(idx)+"+O"+str(idx)+")"

ws = wb['Summary']
ai = getAccountInfo(account)
ws['B5'] = ai[1]
i = 13
for cc in enumerate(ci):
    ws['B'+str(i)] = cc[1]
    c = getCSInfo(cc[1])
    ws['J'+str(i)] = c[0]
    ws['K'+str(i)] = c[1]
    ws['L'+str(i)] = c[2]
    ws['M'+str(i)] = c[3]
    i += 1

i=7
ws = wb['VMs']
mi_sort = sorted(mi,key=getKey)
cs = 0
for x in mi_sort:
    if cs != x[0]:
        i += 1
        cs = x[0]
    ws['A'+str(i)] = x[0]
    ws['B'+str(i)] = x[1]
    ws['C'+str(i)] = x[2]
    ws['D'+str(i)] = x[3]
    ws['E'+str(i)] = x[4]
    ws['F'+str(i)] = x[5]
    ws['G'+str(i)] = x[6]
    ws['H'+str(i)] = x[7]
    ws['I'+str(i)] = x[8]
    ws['J'+str(i)] = x[9]
    ws['L'+str(i)] = "=param!C2*D"+str(i)
    ws['M'+str(i)] = "=param!C3*E"+str(i)
    ws['N'+str(i)] = "=param!C4*F"+str(i)
    ws['O'+str(i)] = "=param!C5*G"+str(i)
    ws['P'+str(i)] = "=param!C6*H"+str(i)
    ws['Q'+str(i)] = "=param!C7*I"+str(i)
    ws['R'+str(i)] = "=param!C8*J"+str(i)
    if x[2].find('Windows') >= 0:
        ws['S'+str(i)] = "=param!B7"
    if x[2].find('Redhat') >= 0:
        ws['T'+str(i)] = "=param!B8"

    i +=1
wb.save(xlsfile)